import secrets
import string
import csv
import os
import time

from django.http import HttpResponse
from django.shortcuts import render
from django.views.generic import CreateView

from apps.eventos.code_gen import CodeGeneration
from apps.eventos.tasks import send_message
from main import settings
from .models import BBCode, Inscricao, Evento, Message, Reward, StoreCode
from .forms import GenerateBBCodeForm, InscricaoForm, RewardsForm, SelectCodesForm, StoreCodesForm
from django.urls import reverse_lazy, reverse

from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser
from .serializers import MessageSerializer
from rest_framework import status
from apps.users.roles import ADMIN, role_required
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage

def detalhe_inscricao(request):
    user = request.user
    i = Inscricao.objects.filter(user=user)
    if i:
        inscricao = i[0]
        separar = inscricao.subevento.local.coordenadas.split()
        coo1 = separar[0][0:15]
        coo2 = separar[1][0:15]
    else:
        i=" "
        separar=" "
        coo1 = " "
        coo2 = " "
    return render(request, 'eventos/inscricao.html', context={'user':user, 'inscricao':inscricao, 'coo1':coo1, 'coo2':coo2})

class NovaInscricao(CreateView):
    form_class = InscricaoForm
    model = Inscricao

    def get(self, request):
        self.user = request.user
        return super().get(request)

    def post(self, request):
        self.user = request.user
        return super().post(request)

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.user.id
        return kwargs

    # def get_context_data(self, **kwargs):
    #     ctx = super().get_context_data(**kwargs)
    #     # ctx['']
    #     return ctx

    def get_success_url(self):
        return reverse('eventos:evento')

def evento(request):
    evento = Evento.objects.last()
    # if request.user.is_authenticated:
    #     ins = Inscricao.objects.filter(user=request.user)
    #     if ins:
    #         return detalhe_inscricao(request)
    #     else:
    #         return render(request, 'eventos/evento.html', context={'evento':evento})
    ins = Inscricao.objects.filter(user=request.user)
    if ins:
        return detalhe_inscricao(request)
    else:
        return render(request, 'eventos/evento.html', context={'evento':evento})
    

def home(request):
    return render(request, 'eventos/home.html',
                    context={'title':'Dia Aberto Bimby®'})

def receitas(request):
    return render(request, 'eventos/receitas.html',
                    context={'title':'Receitas'})

def promos(request):
    return render(request, 'eventos/promos.html',
                    context={'title':'Promoções'})

class BBCodeViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]  # Set permission for this viewset to require authentication (Is set to Basic Auth)
    serializer_class = MessageSerializer  # Set the serializer class for the Message model
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def create(self, request, *args, **kwargs):
        data_dict = dict(request.data)  # Convert the QueryDict object to a dictionary
        for key in data_dict:
            data_dict[key] = str(data_dict[key][0])  # Convert all values to strings
        serializer = self.get_serializer(data=data_dict)  # Get the serializer for the incoming request data
        check_serializer = serializer.is_valid(raise_exception=False)  # Validate the incoming data
        for i in [element for element in data_dict]:
                kwargs[i] = data_dict[i]
        if not check_serializer:  # If validation fails, return a 400 response
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            serializer.save()
            #CodeGeneration.send_message(request,data_dict,**kwargs)
            task = send_message.delay(**kwargs)
            return HttpResponse(b'OK', content_type='text/plain', status=status.HTTP_200_OK)

@login_required
def generate_bbcode(request):
        check_bb = BBCode.objects.filter(user=request.user)
        sms_number = 4242
        REWARD_IMAGES = (
            ('Livro Italia', 'reward_image_1.jpg'),
            ('Livro India', 'reward_image_2.jpg'),
            ('Livro A Cozinhar se Conta uma Historia', 'reward_image_3.jpg'),
            ('Livro ao Sol e ao Por do Sol', 'reward_image_4.jpg'),
            ('Livro Chocolate', 'reward_image_5.jpg'),
            ('Livro Receitas Bimby S e XL', 'reward_image_6.jpg'),
        )

        if check_bb.exists() and Reward.objects.filter(bbcode=check_bb[0]).exists():
            reward = Reward.objects.get(bbcode=check_bb[0]).reward
            context = {
                'result': f'A sua recompensa',
                'reward_type': reward,
                'reward': True,
                'REWARD_IMAGES': REWARD_IMAGES
            }
            return render(request, 'eventos/generate_bbcode.html', context)
        elif check_bb.exists():
            context = {
                'result': f'Envie um SMS para o {sms_number} com o código {check_bb[0].code}',
                'bb_code': True,
            }
            return render(request, 'eventos/generate_bbcode.html', context)
        else:
            form = GenerateBBCodeForm(request.POST)
            if form.is_valid():
                code = form.cleaned_data['code']
                if check_bb.exists():
                    context = {
                        'result': f'Somente um BBCode por cliente... {check_bb[0].code}',
                    }
                    return render(request, 'eventos/generate_bbcode.html', context)
                else:
                    result = CodeGeneration.generate_bb_code(request, code)
                    context = {
                        'result': f'Envie um SMS para o {sms_number} com o código {result}',
                        'bb_code': True
                    }
                    return render(request, 'eventos/generate_bbcode.html', context)
            else:
                form = GenerateBBCodeForm()
            context = {
                'form': form,
            }
            return render(request, 'eventos/generate_bbcode.html', context)

@login_required
@role_required(ADMIN)
def import_store_codes(request):
    from openpyxl import load_workbook
    
    if request.method == 'POST':
        form = StoreCodesForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.cleaned_data['event']
            excel_file = request.FILES['excel_file']

            # Load the Excel file and get the first sheet
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            # Read the codes from the sheet
            codes = [str(cell.value).strip() for cell in sheet['A'] if cell.value and len(str(cell.value).strip()) <= 8]


            # Create StoreCode objects for the codes
            for code in codes:
                StoreCode.objects.create(code=code, event=event)

            # Redirect to the success page
            num_codes = StoreCode.objects.filter(event=event).count()
            return render(request, 'eventos/import_store_codes.html', {'num_codes': num_codes, 'form': form})
    else:
        form = StoreCodesForm()
    return render(request, 'eventos/import_store_codes.html', {'form': form})
 
@login_required
@role_required(ADMIN)
def import_rewards(request):
    from openpyxl import load_workbook
    
    if request.method == 'POST':
        form = RewardsForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.cleaned_data['event']
            excel_file = request.FILES['excel_file']

            # Load the Excel file and get the first sheet
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            # Read the codes from the sheet
            rewards = [str(cell.value).strip() for cell in sheet['A'] if cell.value]

            # Create StoreCode objects for the codes
            for reward in rewards:
                Reward.objects.create(reward=reward, event=event)

            # Redirect to the success page
            num_codes = Reward.objects.all().count()
            return render(request, 'eventos/import_rewards.html', {'num_codes': num_codes, 'form': form})
    else:
        form = StoreCodesForm()
    return render(request, 'eventos/import_rewards.html', {'form': form})


### NOT USED ###

@login_required
@role_required(ADMIN)
def gen_storecodes(request):
    if request.method == 'POST':
        form = StoreCodesForm(request.POST)
        if form.is_valid():
            num_codes = form.cleaned_data['num_codes']
            eventid = form.cleaned_data['event']
            storecodes = CodeGeneration.generate_store_codes(request, num_codes, eventid.id)
            evento = Evento.objects.get(id=eventid.id)
            codes = evento.codes.all()
            unused_codes = codes.filter(is_redeemed=False, picked=False).count()
            return render(request, 'eventos/generete_storecodes.html', {'storecodes': storecodes, 'unused_codes': unused_codes, 'form': form})
    else:
        form = SelectCodesForm()
    return render(request, 'eventos/select_codes.html', {'form': form})

@login_required
@role_required(ADMIN)
def select_storecodes(request):
    if request.method == 'POST':
        form = SelectCodesForm(request.POST)
        if form.is_valid():
            num_codes = form.cleaned_data['num_codes']
            eventid = form.cleaned_data['event']
            selected_codes = CodeGeneration.select_codes(request, num_codes, eventid.id)
            filename = download_codes(request, selected_codes)
            return render(request, 'eventos/select_codes.html', {'selected_codes': selected_codes, 'filename': filename, 'form': form})
    else:
        form = SelectCodesForm()
    return render(request, 'eventos/select_codes.html', {'form': form})

@login_required
@role_required(ADMIN)
def download_codes(request, selected_codes):
    # Generate a unique filename based on the current timestamp
    characters = string.digits + string.ascii_uppercase
    guid = secrets.choice(characters)
    filename = f"selected_codes_{guid}{int(time.time())}.txt"
    filepath = os.path.join(settings.MEDIA_ROOT, filename)

    # Write the selected codes to the file
    try:
        with open(filepath, 'w') as f:
            for code in selected_codes:
                f.write(str(code.code) + '\n')
    except Exception as e:
        # Log or print out the error message
        return HttpResponse("Error generating file.")

    # Get the URL of the file in the media directory
    file_url = default_storage.url(filename)
    return file_url